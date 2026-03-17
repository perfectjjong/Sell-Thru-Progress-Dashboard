"""
Sell-Thru Dashboard Auto Refresh Script
========================================
2026 Raw 데이터 업데이트 후 이 스크립트를 실행하면
Excel(Sell_Thru_Dashboard_Data.xlsx) + HTML(Sell_Thru_Dashboard.html)이 자동 갱신됩니다.

사용법:
  python refresh_dashboard.py
"""

import openpyxl
import pandas as pd
import json
import os
import re
import calendar
import hashlib
import pickle
import time
from collections import Counter
from datetime import datetime, date, timedelta

# ============================================================
# CONFIG
# ============================================================
BASE = os.path.dirname(os.path.abspath(__file__))

FILES = {
    '2024_raw': os.path.join(BASE, '00. 2024', '2024 Sell thru Raw data.xlsx'),
    '2025_raw': os.path.join(BASE, '01. 2025', '2025 Sell thru Raw data.xlsx'),
    '2026_raw': os.path.join(BASE, '03. 2026', '2026 Sell thru Raw data.xlsx'),
    'classification': os.path.join(BASE, '01. Classfication.xlsx'),
}

OUD_DIR = os.path.join(os.path.dirname(BASE), '06. OUD')  # sibling folder

OUTPUT_EXCEL = os.path.join(BASE, 'Sell_Thru_Dashboard_Data.xlsx')
OUTPUT_HTML = os.path.join(BASE, 'Sell_Thru_Dashboard.html')

CACHE_DIR = os.path.join(BASE, '.cache')


def _file_signature(filepath):
    """Return mtime+size string for change detection."""
    if not os.path.exists(filepath):
        return None
    st = os.stat(filepath)
    return f"{st.st_mtime}_{st.st_size}"


def _load_cache(key):
    """Load cached data if exists. Returns (signature, data) or (None, None)."""
    os.makedirs(CACHE_DIR, exist_ok=True)
    cache_file = os.path.join(CACHE_DIR, f'{key}.pkl')
    if os.path.exists(cache_file):
        with open(cache_file, 'rb') as f:
            return pickle.load(f)
    return None, None


def _save_cache(key, signature, data):
    """Save data to cache with file signature."""
    os.makedirs(CACHE_DIR, exist_ok=True)
    cache_file = os.path.join(CACHE_DIR, f'{key}.pkl')
    with open(cache_file, 'wb') as f:
        pickle.dump((signature, data), f)

# Employee Numbers → force SME
SME_EMPLOYEES = {24, 158}

# Manual Team Override (account_id → team) — 2025/2026 only
TEAM_OVERRIDE = {
    1400000008: 'AFS',
    1110000019: 'Projects',
    1500000046: 'AFS',
    1160000004: 'SDA',
    1180000123: 'Projects',
    1180000363: 'Projects',
    1500000024: 'AMC',
    1500000045: 'AFS',
    1500000047: 'AMC',
    1500000058: 'AFS',
}

# Category Mapping
CATEGORY_MAP = {
    'Split Inverter': 'Split Inverter', 'Split on/off': 'Split on/off',
    'Cassette': 'Cassette', 'Concealed': 'Concealed',
    'Convertible (CAC)': 'Convertible (CAC)', 'Free Standing': 'Free Standing',
    'Window': 'Window', 'Multi-V': 'Multi-V', 'AHU': 'AHU',
    'Unitary Package': 'Unitary Package', 'Accessories': 'Accessories',
    'installation': 'installation', 'Others': 'Others',
    'CAC Ducted': 'Concealed', 'Window On/Off': 'Window', 'Window SEEC': 'Window',
}
EXCLUDE_CATS = {'Cooker', 'Dishwasher', 'Laundry', 'SDA', 'RAC', 'FHD', 'Miscellaneous'}

# Quantity ÷ 2 categories (indoor+outdoor set)
HALF_QTY_CATS = {'Cassette', 'Concealed', 'Convertible (CAC)', 'Free Standing', 'Split Inverter', 'Split on/off'}


def normalize_id(val):
    if val is None: return None
    if isinstance(val, (int, float)):
        v = int(val)
        return v if v != 0 else None
    s = str(val).strip()
    if s in ('', 'None', '#N/A'): return None
    try:
        v = int(s)
        return v if v != 0 else None
    except ValueError:
        return s


def map_category(raw_cat):
    if raw_cat in EXCLUDE_CATS: return None
    m = CATEGORY_MAP.get(raw_cat)
    if m: return m
    return 'Others' if raw_cat and raw_cat != '---' else 'Others'


def normalize_emp(val):
    """Normalize employee number to int"""
    if val is None: return None
    try: return int(val)
    except: return None


# ============================================================
# STEP 1: Load Classification
# ============================================================
def load_classification():
    print("1. Loading Classification...")
    wb = openpyxl.load_workbook(FILES['classification'], read_only=True, data_only=True)

    c24 = {}
    for row in wb['2024'].iter_rows(min_row=2, values_only=True):
        r = list(row)
        sap = normalize_id(r[0])
        if sap: c24[sap] = {'name': r[1], 'team': r[5]}

    c25 = {}
    for row in wb['2025'].iter_rows(min_row=2, values_only=True):
        r = list(row)
        cid = normalize_id(r[1])
        if cid: c25[cid] = {'name': r[2], 'team': r[3]}

    wb.close()
    print(f"   2024: {len(c24)} / 2025: {len(c25)} accounts")
    return c24, c25


# ============================================================
# STEP 2: Load Raw Data
# ============================================================
def load_2024(c24):
    """Load 2024 monthly data and distribute evenly across days in each month."""
    print("2. Loading 2024 Raw...")
    wb = openpyxl.load_workbook(FILES['2024_raw'], read_only=True, data_only=True)
    ws = wb['2024_Raw']
    monthly_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        r = list(row)
        cat = map_category(r[3])
        if cat is None: continue
        sap = normalize_id(r[4])
        team = c24.get(sap, {}).get('team', r[21])
        qty = int(r[7]) if r[7] else 0
        if cat in HALF_QTY_CATS: qty = qty // 2
        monthly_rows.append({
            'Year': 2024, 'Month': str(r[1]).zfill(2) if r[1] else '00',
            'Account_ID': sap, 'Account_Name': r[5],
            'Team': team, 'Category': cat,
            'Value': float(r[6] or 0), 'Quantity': qty,
        })
    wb.close()

    # Distribute monthly totals evenly across days
    rows = []
    for mr in monthly_rows:
        mm = int(mr['Month']) if mr['Month'] != '00' else 1
        days_in_month = calendar.monthrange(2024, mm)[1]
        daily_val = mr['Value'] / days_in_month
        daily_qty = mr['Quantity'] / days_in_month
        for d in range(1, days_in_month + 1):
            rows.append({
                'Year': 2024,
                'Month': mr['Month'],
                'Day': f"2024-{mr['Month']}-{str(d).zfill(2)}",
                'Account_ID': mr['Account_ID'], 'Account_Name': mr['Account_Name'],
                'Team': mr['Team'], 'Category': mr['Category'],
                'Value': round(daily_val, 2), 'Quantity': round(daily_qty, 2),
            })
    print(f"   {len(monthly_rows)} monthly rows → {len(rows)} daily rows")
    return rows


def load_raw_2025_2026(path, sheet, c24, c25, year_label):
    print(f"3. Loading {year_label} Raw...")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet]
    rows, skipped = [], 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        r = list(row)
        if len(r) < 101: continue
        cat = map_category(r[77])
        if cat is None:
            skipped += 1
            continue

        inv_date = r[1]
        month = str(inv_date.month).zfill(2) if isinstance(inv_date, datetime) else '00'
        cid = normalize_id(r[38])
        emp_txn = normalize_emp(r[27])  # Sales Employee Number (Transaction)
        classification = r[98]

        # Team logic: Manual override → SME override → Classification file → raw
        if cid in TEAM_OVERRIDE:
            team = TEAM_OVERRIDE[cid]
        elif emp_txn in SME_EMPLOYEES:
            team = 'SME'
        else:
            team = c25.get(cid, {}).get('team')
            if not team: team = c24.get(cid, {}).get('team')
            if not team: team = classification

        qty = int(r[5]) if r[5] else 0
        if cat in HALF_QTY_CATS: qty = qty // 2

        day_str = inv_date.strftime('%Y-%m-%d') if isinstance(inv_date, datetime) else f'{year_label}-{month}-01'
        rows.append({
            'Year': int(year_label), 'Month': month, 'Day': day_str,
            'Account_ID': cid, 'Account_Name': r[39],
            'Team': team, 'Category': cat,
            'Value': float(r[6]) if r[6] else 0, 'Quantity': qty,
        })
    wb.close()
    print(f"   {len(rows)} rows loaded, {skipped} excluded")
    return rows


# ============================================================
# STEP 3: Account Status
# ============================================================
def classify_accounts(df, c24):
    print("4. Classifying Account Status...")
    yearly = df.groupby(['Year', 'Account_ID']).agg(
        Total_Value=('Value', 'sum')).reset_index()

    a24_txn = set(yearly[yearly['Year'] == 2024]['Account_ID'].unique())
    a25_txn = set(yearly[yearly['Year'] == 2025]['Account_ID'].unique())
    a26_txn = set(yearly[yearly['Year'] == 2026]['Account_ID'].unique())
    a24_all = a24_txn | set(c24.keys())

    t24 = yearly[yearly['Year'] == 2024].set_index('Account_ID')['Total_Value'].to_dict()
    t25 = yearly[yearly['Year'] == 2025].set_index('Account_ID')['Total_Value'].to_dict()
    t26 = yearly[yearly['Year'] == 2026].set_index('Account_ID')['Total_Value'].to_dict()

    all_accts = a24_all | a25_txn | a26_txn
    status = {}

    for a in all_accts:
        i24, i25, i26 = a in a24_all, a in a25_txn, a in a26_txn
        v24, v25 = t24.get(a, 0), t25.get(a, 0)
        growth = (v25 - v24) / v24 if v24 > 0 else None

        if i26:
            if not i24 and not i25: s = 'New'
            elif i24 and not i25: s = 'Re-active from 2024'
            elif i25 and not i24: s = 'New_2025'
            elif i25 and i24 and growth is not None and growth >= 0.5: s = 'Re-active_2025'
            else: s = 'Active'
        elif i25:
            if not i24: s = 'New_2025'
            elif growth is not None and growth >= 0.5: s = 'Re-active_2025'
            else: s = 'Active'
        elif i24:
            s = 'Need to re-active'
        else:
            s = 'Unknown'
        status[a] = s

    dist = Counter(status.values())
    print("   " + ", ".join(f"{k}:{v}" for k, v in sorted(dist.items())))
    return status, t24, t25, t26, all_accts


# ============================================================
# STEP 4: Save Excel
# ============================================================
def save_excel(df, status, t24, t25, t26, all_accts, c24, c25):
    print("5. Saving Excel...")
    df['Account_Status'] = df['Account_ID'].map(status)
    df_final = df[['Year', 'Month', 'Day', 'Account_ID', 'Account_Name', 'Team',
                    'Account_Status', 'Category', 'Value', 'Quantity']].copy()
    df_final = df_final.sort_values(['Year', 'Day', 'Account_ID', 'Category'])

    # Account names & teams (latest first)
    names, teams = {}, {}
    for _, r in df_final.sort_values('Year', ascending=False).iterrows():
        a = r['Account_ID']
        if a not in names and r['Account_Name']: names[a] = r['Account_Name']
        if a not in teams and r['Team']: teams[a] = r['Team']
    for a, i in c24.items():
        names.setdefault(a, i.get('name', ''))
        teams.setdefault(a, i.get('team', ''))
    for a, i in c25.items():
        names.setdefault(a, i.get('name', ''))
        teams.setdefault(a, i.get('team', ''))

    master = []
    for a in sorted(all_accts, key=lambda x: str(x)):
        if not a: continue
        v24, v25, v26 = t24.get(a, 0), t25.get(a, 0), t26.get(a, 0)
        master.append({
            'Account_ID': a, 'Account_Name': names.get(a, ''),
            'Team': teams.get(a, ''), 'Account_Status': status.get(a, ''),
            'Value_2024': v24, 'Value_2025': v25, 'Value_2026': v26,
            'Growth_24_25': round((v25 - v24) / v24 * 100, 1) if v24 > 0 else None,
            'Growth_25_26': round((v26 - v25) / v25 * 100, 1) if v25 > 0 else None,
        })

    dm = pd.DataFrame(master)
    with pd.ExcelWriter(OUTPUT_EXCEL, engine='openpyxl') as w:
        df_final.to_excel(w, sheet_name='Transactions', index=False)
        dm.to_excel(w, sheet_name='Account_Master', index=False)

    print(f"   {OUTPUT_EXCEL}")
    print(f"   Transactions: {len(df_final)} / Master: {len(dm)}")
    return df_final, dm


# ============================================================
# STEP 5: Save HTML Dashboard
# ============================================================
def save_html(df_final, dm, oud_data=None):
    print("6. Saving HTML Dashboard...")
    txn = []
    for _, r in df_final.iterrows():
        txn.append([
            int(r['Year']), str(r['Month']).zfill(2),
            str(r['Day']) if pd.notna(r['Day']) else '',
            r['Account_ID'] if pd.notna(r['Account_ID']) else None,
            str(r['Account_Name']) if pd.notna(r['Account_Name']) else '',
            str(r['Team']) if pd.notna(r['Team']) else '',
            str(r['Account_Status']) if pd.notna(r['Account_Status']) else '',
            str(r['Category']) if pd.notna(r['Category']) else '',
            round(float(r['Value']), 2) if pd.notna(r['Value']) else 0,
            round(float(r['Quantity']), 2) if pd.notna(r['Quantity']) else 0
        ])
    master = []
    for _, r in dm.iterrows():
        master.append({
            'id': r['Account_ID'] if pd.notna(r['Account_ID']) else None,
            'name': str(r['Account_Name']) if pd.notna(r['Account_Name']) else '',
            'team': str(r['Team']) if pd.notna(r['Team']) else '',
            'status': str(r['Account_Status']) if pd.notna(r['Account_Status']) else '',
            'v24': round(float(r['Value_2024']), 0) if pd.notna(r['Value_2024']) else 0,
            'v25': round(float(r['Value_2025']), 0) if pd.notna(r['Value_2025']) else 0,
            'v26': round(float(r['Value_2026']), 0) if pd.notna(r['Value_2026']) else 0,
        })

    # Build OUD JSON
    oud_json = {'current': None, 'prev': None}
    if oud_data:
        for key in ['current', 'prev']:
            if key in oud_data and oud_data[key]:
                d = oud_data[key]
                accts = {}
                for aid, info in d['accounts'].items():
                    accts[str(aid)] = {
                        'nm': info['name'],
                        'v': round(info['value'], 0),
                        'q': round(info['qty'], 0),
                    }
                oud_json[key] = {
                    'date': d['date'],
                    'tv': round(d['total_value'], 0),
                    'tq': round(d['total_qty'], 0),
                    'accts': accts,
                }

    json_str = json.dumps({'txn': txn, 'master': master, 'oud': oud_json}, ensure_ascii=False, separators=(',', ':'))

    with open(OUTPUT_HTML, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    # Replace the RAW data line
    new_lines = []
    replaced = False
    for line in lines:
        if line.strip().startswith('const RAW') and not replaced:
            new_lines.append('const RAW = ' + json_str + ';\n')
            replaced = True
        else:
            new_lines.append(line)

    if not replaced:
        print("   WARNING: Could not find RAW data line to replace!")

    html = ''.join(new_lines)

    # Update date in header
    today = datetime.now().strftime('%Y-%m-%d')
    html = re.sub(r'Updated</span><span class="val">\d{4}-\d{2}-\d{2}',
                  f'Updated</span><span class="val">{today}', html)

    with open(OUTPUT_HTML, 'w', encoding='utf-8') as f:
        f.write(html)

    print(f"   {OUTPUT_HTML} ({len(html):,} bytes)")


# ============================================================
# STEP 6: Merge Duplicate Account Names
# ============================================================
def merge_duplicate_accounts(df):
    """
    동일 Account_Name에 복수 Account_ID가 있는 경우,
    총 매출(Value) 기준 가장 큰 ID를 대표 ID로 선정하고 나머지를 합산.
    """
    print("   Merging duplicate Account Names...")

    # Build name → {id: total_value}
    name_ids = {}
    for _, r in df.iterrows():
        nm = str(r['Account_Name']).strip() if r['Account_Name'] else ''
        if not nm:
            continue
        aid = r['Account_ID']
        val = float(r['Value']) if r['Value'] else 0
        if nm not in name_ids:
            name_ids[nm] = {}
        name_ids[nm][aid] = name_ids[nm].get(aid, 0) + val

    # Find names with multiple IDs → pick representative (highest total value)
    id_remap = {}  # old_id → representative_id
    merge_count = 0
    for nm, ids_vals in name_ids.items():
        if len(ids_vals) <= 1:
            continue
        # Representative = ID with highest total value
        rep_id = max(ids_vals, key=lambda x: ids_vals[x])
        for aid in ids_vals:
            if aid != rep_id:
                id_remap[aid] = rep_id
        merge_count += 1

    if id_remap:
        df['Account_ID'] = df['Account_ID'].map(lambda x: id_remap.get(x, x))
        print(f"   → {merge_count} account names merged ({len(id_remap)} IDs remapped)")
    else:
        print("   → No duplicates found")

    return df


# ============================================================
# STEP 7: Load OUD Data
# ============================================================
def parse_oud_date(filename):
    """Extract date from filename like '14-MAR-2026 HVAC.xlsx' → '2026-03-14'"""
    import locale
    m = re.match(r'(\d{2})-([A-Z]{3})-(\d{4})', filename)
    if not m:
        return None
    day, mon_str, year = m.group(1), m.group(2), m.group(3)
    months = {'JAN':'01','FEB':'02','MAR':'03','APR':'04','MAY':'05','JUN':'06',
              'JUL':'07','AUG':'08','SEP':'09','OCT':'10','NOV':'11','DEC':'12'}
    mm = months.get(mon_str)
    if not mm:
        return None
    return f"{year}-{mm}-{day}"


def load_oud(id_remap=None):
    """
    Load OUD files from 06. OUD folder.
    Returns: {
        'current': {'date': '2026-03-14', 'accounts': {account_id: {name, value, qty}}, 'total_value', 'total_qty'},
        'prev':    {'date': '2026-03-07', 'accounts': {account_id: {name, value, qty}}, 'total_value', 'total_qty'},
    }
    """
    print("7. Loading OUD Data...")

    if not os.path.isdir(OUD_DIR):
        print(f"   OUD folder not found: {OUD_DIR}")
        return None

    # Find all OUD xlsx files and sort by date
    oud_files = []
    for f in os.listdir(OUD_DIR):
        if f.endswith('.xlsx') and not f.startswith('~$'):
            dt = parse_oud_date(f)
            if dt:
                oud_files.append((dt, f))
    oud_files.sort(key=lambda x: x[0])

    if not oud_files:
        print("   No OUD files found")
        return None

    print(f"   Found {len(oud_files)} OUD files: {[f[1] for f in oud_files]}")

    # OUD Group → Sell-Thru category mapping (OUD names are truncated)
    OUD_GROUP_MAP = {
        'Split Inve': 'Split Inverter',
        'Split on/o': 'Split on/off',
        'Cassette':   'Cassette',
        'Concealed':  'Concealed',
        'Free Stand': 'Free Standing',
        'CAC Ducted': 'Convertible (CAC)',
        'Window SEE': 'Window',
        'Window SEEC':'Window',
        'AHU':        'AHU',
        'Multi-V':    'Multi-V',
        'Unitary Pa': 'Unitary Package',
        'Accessorie': 'Accessories',
    }

    def read_oud_file(filepath):
        """Read a single OUD file → {account_id: {name, value, qty}}"""
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

        # Find header row (row with 'Customer Name')
        header_idx = None
        for i, r in enumerate(rows):
            if r and r[0] == 'Customer Name':
                header_idx = i
                break
        if header_idx is None:
            return {}

        headers = [str(h).strip() if h else '' for h in rows[header_idx]]

        # Find column indices dynamically
        def find_col(name):
            for i, h in enumerate(headers):
                if h == name:
                    return i
            return None

        col_name = find_col('Customer Name')
        col_acc = find_col('Dealer ACC No') or find_col('Column1')
        col_rqty = find_col('R-Qty')
        col_tval = find_col('Total Value')
        col_group = find_col('Group')

        accounts = {}
        for r in rows[header_idx + 1:]:
            if not r or not r[col_name]:
                continue
            name = str(r[col_name]).strip()
            if 'Total' in name or name == 'Grand Total':
                continue

            acc_id = normalize_id(r[col_acc]) if col_acc is not None else None
            # Apply id_remap if available
            if id_remap and acc_id in id_remap:
                acc_id = id_remap[acc_id]

            rqty = float(r[col_rqty]) if col_rqty is not None and r[col_rqty] else 0
            tval = float(r[col_tval]) if col_tval is not None and r[col_tval] else 0

            # Apply HALF_QTY_CATS logic based on Group column
            if col_group is not None and r[col_group]:
                grp = str(r[col_group]).strip()
                mapped_cat = OUD_GROUP_MAP.get(grp, grp)
                if mapped_cat in HALF_QTY_CATS:
                    rqty = rqty / 2

            if acc_id not in accounts:
                accounts[acc_id] = {'name': name, 'value': 0, 'qty': 0}
            accounts[acc_id]['value'] += tval
            accounts[acc_id]['qty'] += rqty

        return accounts

    result = {}

    # Current = latest file
    cur_date, cur_file = oud_files[-1]
    cur_accounts = read_oud_file(os.path.join(OUD_DIR, cur_file))
    result['current'] = {
        'date': cur_date,
        'accounts': cur_accounts,
        'total_value': sum(a['value'] for a in cur_accounts.values()),
        'total_qty': sum(a['qty'] for a in cur_accounts.values()),
    }
    print(f"   Current: {cur_file} → {len(cur_accounts)} accounts, Value={result['current']['total_value']:,.0f}")

    # Previous = second to last file (if exists)
    if len(oud_files) >= 2:
        prev_date, prev_file = oud_files[-2]
        prev_accounts = read_oud_file(os.path.join(OUD_DIR, prev_file))
        result['prev'] = {
            'date': prev_date,
            'accounts': prev_accounts,
            'total_value': sum(a['value'] for a in prev_accounts.values()),
            'total_qty': sum(a['qty'] for a in prev_accounts.values()),
        }
        print(f"   Previous: {prev_file} → {len(prev_accounts)} accounts, Value={result['prev']['total_value']:,.0f}")
        delta_v = result['current']['total_value'] - result['prev']['total_value']
        delta_q = result['current']['total_qty'] - result['prev']['total_qty']
        print(f"   Delta: Value={delta_v:+,.0f}, Qty={delta_q:+,.0f}")

    return result


# ============================================================
# MAIN
# ============================================================
def main():
    t_start = time.time()
    print("=" * 50)
    print("Sell-Thru Dashboard Refresh")
    print("=" * 50)

    # --- Classification (needed for all years) ---
    cls_sig = _file_signature(FILES['classification'])
    cached_cls_sig, cached_cls = _load_cache('classification')
    if cached_cls_sig == cls_sig and cached_cls is not None:
        c24, c25 = cached_cls
        print(f"1. Classification → CACHED ({len(c24)}/{len(c25)} accounts)")
    else:
        c24, c25 = load_classification()
        _save_cache('classification', cls_sig, (c24, c25))

    # --- 2024 Data (cache by source file + classification) ---
    sig_24 = f"{_file_signature(FILES['2024_raw'])}|{cls_sig}"
    cached_sig_24, cached_r24 = _load_cache('raw_2024')
    if cached_sig_24 == sig_24 and cached_r24 is not None:
        r24 = cached_r24
        print(f"2. 2024 Raw → CACHED ({len(r24)} daily rows)")
    else:
        r24 = load_2024(c24)
        _save_cache('raw_2024', sig_24, r24)

    # Override hash (cache bust when TEAM_OVERRIDE or SME_EMPLOYEES change)
    ovr_hash = str(sorted(TEAM_OVERRIDE.items())) + str(sorted(SME_EMPLOYEES))

    # --- 2025 Data ---
    sig_25 = f"{_file_signature(FILES['2025_raw'])}|{cls_sig}|{ovr_hash}"
    cached_sig_25, cached_r25 = _load_cache('raw_2025')
    if cached_sig_25 == sig_25 and cached_r25 is not None:
        r25 = cached_r25
        print(f"3. 2025 Raw → CACHED ({len(r25)} rows)")
    else:
        r25 = load_raw_2025_2026(FILES['2025_raw'], 'RAW', c24, c25, '2025')
        _save_cache('raw_2025', sig_25, r25)

    # --- 2026 Data (always check, most likely to change) ---
    sig_26 = f"{_file_signature(FILES['2026_raw'])}|{cls_sig}|{ovr_hash}"
    cached_sig_26, cached_r26 = _load_cache('raw_2026')
    if cached_sig_26 == sig_26 and cached_r26 is not None:
        r26 = cached_r26
        print(f"3. 2026 Raw → CACHED ({len(r26)} rows)")
    else:
        r26 = load_raw_2025_2026(FILES['2026_raw'], 'Raw', c24, c25, '2026')
        _save_cache('raw_2026', sig_26, r26)

    df = pd.DataFrame(r24 + r25 + r26)

    # ---- Merge duplicate Account Names → single representative ID ----
    df = merge_duplicate_accounts(df)

    df_agg = df.groupby(['Year', 'Month', 'Day', 'Account_ID', 'Account_Name', 'Team', 'Category']).agg(
        Value=('Value', 'sum'), Quantity=('Quantity', 'sum')).reset_index()

    status, t24, t25, t26, all_accts = classify_accounts(df_agg, c24)
    df_final, dm = save_excel(df_agg, status, t24, t25, t26, all_accts, c24, c25)

    # Build id_remap for OUD (same logic as merge_duplicate_accounts)
    name_ids = {}
    for _, r in df.iterrows():
        nm = str(r['Account_Name']).strip() if r['Account_Name'] else ''
        if not nm: continue
        aid = r['Account_ID']
        val = float(r['Value']) if r['Value'] else 0
        if nm not in name_ids: name_ids[nm] = {}
        name_ids[nm][aid] = name_ids[nm].get(aid, 0) + val
    id_remap = {}
    for nm, ids_vals in name_ids.items():
        if len(ids_vals) <= 1: continue
        rep_id = max(ids_vals, key=lambda x: ids_vals[x])
        for aid in ids_vals:
            if aid != rep_id: id_remap[aid] = rep_id

    oud_data = load_oud(id_remap)
    save_html(df_final, dm, oud_data)

    elapsed = time.time() - t_start
    print("\n" + "=" * 50)
    print(f"DONE! ({elapsed:.1f}s)")
    for yr in [2024, 2025, 2026]:
        sub = df_final[df_final['Year'] == yr]
        print(f"  {yr}: {sub['Account_ID'].nunique()} accounts / Value={sub['Value'].sum():,.0f} / Qty={sub['Quantity'].sum():,.0f}")

    # SME check
    sme = df_final[df_final['Team'] == 'SME']
    print(f"\n  SME transactions: {len(sme)} rows, {sme['Account_ID'].nunique()} accounts")
    print("=" * 50)


if __name__ == '__main__':
    main()
